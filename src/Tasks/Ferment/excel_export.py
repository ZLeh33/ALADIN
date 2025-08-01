from openpyxl import Workbook
import os
from pathlib import Path
from io import BytesIO

def export_to_excel(filename, t_combined, y_combined, cum_feeding):
    y_combined = y_combined.T
    
    # Attempt to get the user's Downloads folder
    downloads_path = Path.home() / "Downloads"
    
    # Check if Downloads path exists; if not, create it (optional)
    if not downloads_path.exists():
        print("Downloads folder not found.")
        return
    
    # Combine the Downloads folder path with the provided filename
    file_path = downloads_path / filename
    
    # If the file already exists, remove it
    if os.path.exists(file_path):
        os.remove(file_path)
    
    wb = Workbook()
    ws = wb.active
    
    caption = ['t', 'Biotrockenmasse (cx)', 'Konz. Substrat 1 (cs1)', 'Konz. Substrat 2 (cs2)',
               'Konz. Produkt (cp)', 'c_O2 (Lösung)', 'c_O2 (Abluft)', 'c_CO2 (Abluft)',
               'kumulatives Feeding Substrat 1']
    
    # Write headers
    for i, header in enumerate(caption):
        ws.cell(row=1, column=i + 1, value=header)
    
    # Write data
    for i in range(len(t_combined)):
        row_data = [t_combined[i]] + [y_combined[j, i] for j in range(y_combined.shape[0])] + [cum_feeding[i]]
        for j, value in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=value)
    
    # Save the file in the correct Downloads folder
    wb.save(file_path)

    #print(f"File saved to: {file_path}")


    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    
    return file_obj

