
from openpyxl import Workbook
import os


def export_to_excel(filename, t_combined, y_combined, cum_feeding):
	y_combined=y_combined.T
	if os.path.exists(filename):
		os.remove(filename)
	
	wb = Workbook()
	ws = wb.active
	
	caption = ['t', 'Biotrockenmasse (cx)', 'Konz. Substrat 1 (cs1)', 'Konz. Substrat 2 (cs2)',
            'Konz. Produkt (cp)', 'c_O2 (Lösung)', 'c_O2 (Abluft)', 'c_CO2 (Abluft)',
            'kumulatives Feeding Substrat 1']
	
	for i, header in enumerate(caption):
		ws.cell(row=1, column=i + 1, value=header)
		
	for i in range(len(t_combined)):
		row_data = [t_combined[i]] + [y_combined[j, i] for j in range(y_combined.shape[0])] + [cum_feeding[i]]
		for j, value in enumerate(row_data):
			ws.cell(row=i + 2, column=j + 1, value=value)
	
	
	wb.save(filename)